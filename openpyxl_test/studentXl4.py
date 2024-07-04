from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active
ws.title = "수강생_정보"

# 열로 입력할 데이터를 리스트로 만들어 'data'에 저장
data = [ '이철수', '김미소', '최학준' ]

# for문으로 'A'열의 각 셀에 순서대로 접근하여 데이터를 입력
for i, value in enumerate(data):
    ws.cell(row=i+1, column=1, value=value)
    print(i , ' ', value)
    
    
from openpyxl.utils.exceptions import InvalidFileException
'''
try :
    wb = load_workbook()
except (FileNotFoundError, InvalidFileException) as e:
    print(f"파일을 열 수 없습니다: {e}")
except Exception as e:
    print(f"알 수 없는 예외 발생: {e}")
   
''' 
# Exception 으로 죄다 잡지만 파일 에러만 못잡는다 

wb = load_workbook()

print('위에서 오류가 생기면 안찍힌다')
#wb.save("수강생_리스트.xlsx")
